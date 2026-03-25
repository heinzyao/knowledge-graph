"""
FastAPI routes: graph data, node preview, tag list, WebSocket updates.
Module-level globals (graph_builder, base_dir, ws_connections) are
injected by main.py at startup.
"""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from fastapi import APIRouter, HTTPException, Query, WebSocket, WebSocketDisconnect

if TYPE_CHECKING:
    from core.graph_builder import GraphBuilder

router = APIRouter()

# Injected by main.py
graph_builder: GraphBuilder = None  # type: ignore[assignment]
base_dir: Path = None               # type: ignore[assignment]
ws_connections: list[WebSocket] = []


# ------------------------------------------------------------------
# REST endpoints
# ------------------------------------------------------------------

@router.get("/api/graph")
async def get_graph(include_implicit: bool = Query(default=False)) -> dict[str, Any]:
    return graph_builder.to_d3_format(include_implicit=include_implicit)


@router.get("/api/tags")
async def get_tags() -> list[str]:
    return graph_builder.get_all_tags()


@router.get("/api/node")
async def get_node(
    id: str = Query(..., description="Node ID (relative path)"),
    sheet: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=5000),
) -> dict[str, Any]:
    node = graph_builder.get_node_data(id)
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")

    file_path = base_dir / id
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        content_sheets = [s for s in wb.sheetnames if s != "Meta"]

        # Choose which sheet to preview
        target = sheet if (sheet and sheet in wb.sheetnames) else (content_sheets[0] if content_sheets else None)

        preview: dict[str, list] = {}
        total_rows = 0

        if target:
            ws = wb[target]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= limit:
                    break
                rows.append([("" if c is None else str(c)) for c in row])
            preview[target] = rows
            total_rows = ws.max_row or 0

        wb.close()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {
        "meta": {
            "id": node["id"],
            "title": node["title"],
            "tags": node["tags"],
            "links": node["links"],
            "description": node["description"],
        },
        "sheet_names": node["sheet_names"],
        "active_sheet": target,
        "preview": preview,
        "total_rows": total_rows,
        "limit": limit,
    }


@router.post("/api/refresh")
async def refresh_graph() -> dict[str, Any]:
    """Force a full rescan of the data directory."""
    graph_builder.build(include_implicit=True)
    await _broadcast({"type": "refresh"})
    return {"status": "ok", "node_count": graph_builder.node_count}


# ------------------------------------------------------------------
# WebSocket
# ------------------------------------------------------------------

@router.websocket("/ws/updates")
async def ws_updates(websocket: WebSocket) -> None:
    await websocket.accept()
    ws_connections.append(websocket)
    try:
        while True:
            # Keep connection alive; actual messages are pushed from the server
            await websocket.receive_text()
    except WebSocketDisconnect:
        pass
    finally:
        if websocket in ws_connections:
            ws_connections.remove(websocket)


async def _broadcast(message: dict[str, Any]) -> None:
    dead: list[WebSocket] = []
    for ws in list(ws_connections):
        try:
            await ws.send_json(message)
        except Exception:
            dead.append(ws)
    for ws in dead:
        if ws in ws_connections:
            ws_connections.remove(ws)


# Expose broadcast so main.py can call it from the watcher callback
broadcast = _broadcast
