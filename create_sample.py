"""
建立範例 Excel 檔案供測試。
執行方式：uv run python create_sample.py
"""
import openpyxl
from pathlib import Path


def meta_sheet(wb, title: str, tags: str, links: str, description: str):
    ws = wb.active
    ws.title = "Meta"
    ws.append(["title",       title])
    ws.append(["tags",        tags])
    ws.append(["links",       links])
    ws.append(["description", description])
    return ws


def save(wb, folder: Path, name: str):
    wb.save(folder / f"{name}.xlsx")
    print(f"  建立：{name}.xlsx")


def main():
    folder = Path(__file__).parent / "sample_data"
    folder.mkdir(exist_ok=True)

    # ── 1. 專案計畫書 ──────────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(wb, "專案計畫書", "專案, 2026, 重要", "技術規格, 客戶需求", "2026年 Q1 主要專案整體計畫")

    ws = wb.create_sheet("任務清單")
    ws.append(["任務", "負責人", "截止日期", "狀態"])
    ws.append(["需求分析", "Alice", "2026-02-01", "完成"])
    ws.append(["系統設計", "Bob",   "2026-02-15", "進行中"])
    ws.append(["後端開發", "Charlie","2026-03-15","待開始"])
    ws.append(["前端開發", "Diana", "2026-03-31", "待開始"])
    ws.append(["測試",     "Eve",   "2026-04-10", "待開始"])

    ws2 = wb.create_sheet("預算")
    ws2.append(["項目", "預算 (萬)", "實際 (萬)", "差異"])
    ws2.append(["人力", 80, 72, -8])
    ws2.append(["基礎設施", 15, 18, 3])
    ws2.append(["其他", 5, 4, -1])

    save(wb, folder, "專案計畫書")

    # ── 2. 技術規格 ────────────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(wb, "技術規格", "技術, 後端", "資料庫設計, 專案計畫書", "系統技術架構與 API 規格")

    ws = wb.create_sheet("API 清單")
    ws.append(["端點", "方法", "說明", "狀態"])
    ws.append(["/api/users",    "GET",    "取得用戶列表", "完成"])
    ws.append(["/api/users",    "POST",   "建立用戶",     "完成"])
    ws.append(["/api/products", "GET",    "取得商品列表", "進行中"])
    ws.append(["/api/orders",   "POST",   "建立訂單",     "待開始"])

    ws2 = wb.create_sheet("技術棧")
    ws2.append(["層級", "技術", "版本", "備註"])
    ws2.append(["後端",   "FastAPI",    "0.115", "主框架"])
    ws2.append(["資料庫", "PostgreSQL", "16",    "主資料庫"])
    ws2.append(["快取",   "Redis",      "7.2",   "Session & 快取"])
    ws2.append(["前端",   "React",      "18",    "SPA"])

    save(wb, folder, "技術規格")

    # ── 3. 客戶需求 ────────────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(wb, "客戶需求", "需求, 客戶", "專案計畫書", "客戶訪談與功能需求整理")

    ws = wb.create_sheet("功能需求")
    ws.append(["需求編號", "功能", "優先級", "來源", "狀態"])
    ws.append(["REQ-001", "用戶登入/登出",  "高", "客戶A", "確認"])
    ws.append(["REQ-002", "商品搜尋",       "高", "客戶A", "確認"])
    ws.append(["REQ-003", "購物車",         "高", "客戶B", "確認"])
    ws.append(["REQ-004", "訂單追蹤",       "中", "客戶A", "討論中"])
    ws.append(["REQ-005", "推薦系統",       "低", "內部",  "待評估"])

    save(wb, folder, "客戶需求")

    # ── 4. 資料庫設計 ──────────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(wb, "資料庫設計", "技術, 資料庫", "技術規格", "資料表結構與 ER Diagram 說明")

    ws = wb.create_sheet("資料表")
    ws.append(["資料表", "欄位", "型別", "說明"])
    ws.append(["users",    "id",         "UUID",      "主鍵"])
    ws.append(["users",    "email",      "VARCHAR",   "唯一索引"])
    ws.append(["users",    "created_at", "TIMESTAMP", "建立時間"])
    ws.append(["products", "id",         "UUID",      "主鍵"])
    ws.append(["products", "name",       "VARCHAR",   "商品名稱"])
    ws.append(["products", "price",      "DECIMAL",   "售價"])
    ws.append(["orders",   "id",         "UUID",      "主鍵"])
    ws.append(["orders",   "user_id",    "UUID",      "FK → users"])

    save(wb, folder, "資料庫設計")

    # ── 5. 團隊成員 ────────────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(wb, "團隊成員", "人員, 組織", "專案計畫書", "專案團隊組成與職責")

    ws = wb.create_sheet("成員清單")
    ws.append(["姓名", "角色", "專長", "Email"])
    ws.append(["Alice",   "PM",       "專案管理, 需求分析",  "alice@example.com"])
    ws.append(["Bob",     "架構師",   "系統設計, 後端",      "bob@example.com"])
    ws.append(["Charlie", "後端工程師","Python, FastAPI",    "charlie@example.com"])
    ws.append(["Diana",   "前端工程師","React, TypeScript",  "diana@example.com"])
    ws.append(["Eve",     "QA",       "測試自動化, Selenium","eve@example.com"])

    save(wb, folder, "團隊成員")

    # ── 6. 測試計畫（無 Meta → fallback 節點）────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "測試案例"
    ws.append(["案例編號", "測試項目", "預期結果", "狀態"])
    ws.append(["TC-001", "登入功能",     "成功登入並跳轉",     "通過"])
    ws.append(["TC-002", "非法登入",     "顯示錯誤訊息",       "通過"])
    ws.append(["TC-003", "商品搜尋",     "回傳符合結果",       "通過"])
    ws.append(["TC-004", "購物車加入",   "數量正確更新",       "失敗"])
    ws.append(["TC-005", "訂單送出",     "訂單編號產生",       "待測"])
    # 隱式連結：儲存格含有其他節點名稱
    ws2 = wb.create_sheet("關聯說明")
    ws2.append(["此測試計畫依據以下文件："])
    ws2.append(["技術規格"])
    ws2.append(["客戶需求"])

    save(wb, folder, "測試計畫")

    print(f"\n完成！共建立 6 個範例 Excel 於 {folder}")
    print("執行：uv run python main.py  即可啟動")


if __name__ == "__main__":
    main()
