"""
Scan a directory for .xlsx files and extract node metadata.
Each file can have an optional 'Meta' sheet with key-value pairs:
  title, tags (comma-sep), links (comma-sep filenames), description
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def node_id_from_path(file_path: Path, base_dir: Path) -> str:
    """Stable unique ID: relative path with forward slashes."""
    return str(file_path.relative_to(base_dir)).replace("\\", "/")


def parse_excel(file_path: Path, base_dir: Path) -> dict[str, Any]:
    """Parse a single .xlsx file and return node data dict."""
    nid = node_id_from_path(file_path, base_dir)

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        return {
            "id": nid,
            "title": file_path.stem,
            "tags": [],
            "links": [],
            "description": "",
            "sheet_names": [],
            "has_meta": False,
            "error": str(exc),
        }

    sheet_names = wb.sheetnames
    meta: dict[str, str] = {}

    if "Meta" in sheet_names:
        ws = wb["Meta"]
        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            if row and row[0] is not None and row[1] is not None:
                key = str(row[0]).strip().lower()
                val = str(row[1]).strip()
                if key:
                    meta[key] = val

    wb.close()

    content_sheets = [s for s in sheet_names if s != "Meta"]
    has_meta = "Meta" in sheet_names

    title = meta.get("title", file_path.stem)
    tags_raw = meta.get("tags", "")
    tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []
    links_raw = meta.get("links", "")
    links = [lnk.strip() for lnk in links_raw.split(",") if lnk.strip()] if links_raw else []
    description = meta.get("description", "")

    return {
        "id": nid,
        "title": title,
        "tags": tags,
        "links": links,      # explicit link targets (title or stem)
        "description": description,
        "sheet_names": content_sheets,
        "has_meta": has_meta,
    }


def scan_directory(folder: Path) -> list[dict[str, Any]]:
    """Scan folder recursively for .xlsx files. Skip temp files (~$...)."""
    nodes = []
    for xlsx in sorted(folder.rglob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        nodes.append(parse_excel(xlsx, folder))
    return nodes


def scan_implicit_links(
    file_path: Path,
    base_dir: Path,
    known_names: set[str],
) -> list[str]:
    """
    Scan every cell in the file (excluding Meta sheet) for values that
    exactly match a known node title or filename stem.
    Returns list of matched names (may contain duplicates if multiple cells match).
    """
    matches: set[str] = set()
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name == "Meta":
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        val = cell.strip()
                        if val in known_names:
                            matches.add(val)
        wb.close()
    except Exception:
        pass
    return list(matches)
